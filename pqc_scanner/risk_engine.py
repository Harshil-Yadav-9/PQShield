import json, os, sys, re
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCORE = {"Low": 10, "Acceptable": 5, "Medium": -2, "High": -6, "Critical": -10}

SEV_STYLE = {
    "Critical":   ("FFFFC7CE", "FF9C0006"),
    "High":       ("FFFFC7CE", "FF9C0006"),
    "Medium":     ("FFFFEB9C", "FF9C5700"),
    "Acceptable": ("FFC6EFCE", "FF006100"),
    "Low":        ("FFC6EFCE", "FF006100"),
    "N/A":        ("00000000", "00000000"),
}

C = {
    "SEC_BG": "FFFFFFFF",
    "SEC_FG": "FF000000",
    "COL_BG": "FFFFFFFF",
    "COL_FG": "FF000000",
    "TITLE_BG": "FFFFFFFF",
    "TITLE_FG": "FF000000"
}

_S = Side(border_style="thin", color="FFCCCCCC")
BRD = Border(left=_S, right=_S, top=_S, bottom=_S)
WRP = Alignment(horizontal="left", vertical="center", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center", wrap_text=True)

def _f(h): return PatternFill("solid", fgColor=h) if h != "00000000" else PatternFill()
def _font(c="FF000000", b=False, s=10, u=None): return Font(name="Times New Roman", size=s, bold=b, color=c, underline=u)
def _title_font(c="FF000000", b=False, s=10, u=None): return Font(name="Times New Roman", size=s, bold=b, color=c, underline=u)
def _j(v): return " | ".join(str(x) for x in v) if isinstance(v, list) else (str(v) if v else "")
def _cert_key_type(sig_alg, algo_name):
    s = str(sig_alg or "").upper()
    a = str(algo_name or "").upper()
    if any(x in s for x in ("MLDSA", "ML-DSA", "DILITHIUM")) or any(x in a for x in ("MLDSA", "DILITHIUM")):
        return "Dilithium/Falcon/SPHINCS (ML/SLH DSA)"
    if any(x in s for x in ("SLHDSA", "SLH-DSA", "SPHINCS")) or any(x in a for x in ("SLHDSA", "SPHINCS")):
        return "Dilithium/Falcon/SPHINCS (ML/SLH DSA)"
    if any(x in s for x in ("FALCON", "FNDSA", "FN-DSA")) or any(x in a for x in ("FALCON", "FNDSA", "FN-DSA")):
        return "Dilithium/Falcon/SPHINCS (ML/SLH DSA)"
    if "ED25519" in s or "ED448" in s or "ED25519" in a or "ED448" in a:
        return "EdDSA"
    if "ECDSA" in s or ("EC" in a and "RSA" not in a and "ECDSA" not in a):
        return "ECDSA"
    if "RSA" in s or "RSA" in a:
        return "RSA"
    if "DSA" in s or "DSA" in a:
        return "DSA"
    return "Anything else"

def _cert_size_label(algo_name, key_size):
    a = str(algo_name or "").upper()
    if any(x in a for x in ("MLKEM", "KYBER", "DILITHIUM", "SLHDSA", "FALCON", "SPHINCS")):
        return "ANYTHING IN PQC"
    if "RSA-" in a:
        m = re.search(r"RSA[-_ ]?(\d+)", a)
        if m:
            stated = int(m.group(1))
            k = key_size if key_size and key_size != stated else stated
        else:
            k = key_size
        if k >= 4096:
            return "RSA 4096 & above"
        if k >= 3072:
            return "RSA 3072"
        if k >= 2048:
            return "RSA 2048"
        if k == 1024:
            return "RSA 1024"
        if k > 0:
            return f"RSA {k}"
        return "RSA 1024 & LESSER"
    if "RSA" in a:
        if key_size >= 4096:
            return "RSA 4096 & above"
        if key_size >= 3072:
            return "RSA 3072"
        if key_size >= 2048:
            return "RSA 2048"
        if key_size == 1024:
            return "RSA 1024"
        if key_size > 0:
            return f"RSA {key_size}"
        return "RSA 1024 & LESSER"
    if "ECDSA-" in a or "EC-" in a:
        m = re.search(r"(?:ECDSA|EC)[-_ ]?(\d+)", a)
        k = int(m.group(1)) if m else key_size
        if k >= 384:
            return "secp384r1 & above"
        if k == 256:
            return "secp256r1"
        if k == 224:
            return "secp224r1"
        if k > 0:
            return f"EC {k}"
        return "ANYTHING ELSE"
    if "SECP384" in a or key_size >= 384:
        return "secp384r1 & above"
    if "SECP256" in a or key_size == 256:
        return "secp256r1"
    if "SECP224" in a or key_size == 224:
        return "secp224r1"
    if "ED25519" in a:
        return "Ed25519"
    if "ED448" in a:
        return "Ed448"
    return "ANYTHING ELSE"

def _cert_algo_display(sig_alg, algo_name, key_size):
    a = str(algo_name or "").strip()
    s = str(sig_alg or "").upper()
    if a:
        if "RSA" in a:
            if key_size and key_size > 0:
                return f"RSA-{key_size}"
            return "RSA"
        if any(x in a.upper() for x in ("ECDSA", "ED25519", "ED448", "DILITHIUM", "FALCON", "SPHINCS", "KYBER", "SLHDSA")):
            return a
    if "RSA" in s:
        return f"RSA-{key_size}" if key_size and key_size > 0 else "RSA"
    if "ECDSA" in s:
        return f"ECDSA-{key_size}" if key_size and key_size > 0 else "ECDSA"
    if "ED25519" in s:
        return "Ed25519"
    if "ED448" in s:
        return "Ed448"
    return a or sig_alg or "Unknown"

def _normalize_cert_key_size(algo_name, key_size):
    assert key_size is not None
    k = int(key_size) if isinstance(key_size, (int, float)) else 0
    if k <= 0:
        return 0
    a = str(algo_name or "").upper()
    if "RSA-" in a:
        m = re.search(r"RSA[-_ ]?(\d+)", a)
        if m:
            stated = int(m.group(1))
            if k == stated:
                return k
            if k * 8 == stated:
                return k * 8
    if "RSA" in a and k < 128:
        return k * 8
    if any(x in a for x in ("ECDSA", "EC")) and k in (28, 32, 48, 66, 72, 96):
        return k * 8
    if any(x in a for x in ("ED25519", "ED448")) and k in (32, 57):
        return k * 8
    return k

def _signature_label(sig_alg):
    s = str(sig_alg or "").upper()
    if any(x in s for x in ("MLDSA", "ML-DSA", "SLHDSA", "SLH-DSA", "FALCON", "FNDSA", "FN-DSA")):
        return "ML-DSA, SLH DSA"
    if any(x in s for x in ("ECDSA",)) and any(x in s for x in ("SHA256", "SHA384")):
        return "ECDSA + SHA256/384"
    if "RSA" in s and any(x in s for x in ("SHA256", "SHA384")):
        return "RSA + SHA256/384"
    if any(x in s for x in ("SHA1", "SHA-1", "S224")) and any(x in s for x in ("ECDSA", "RSA", "DSA")):
        return "ECDSA/RSA + SHA-1 or (S224)"
    if "SHA224" in s:
        return "SHA 224"
    if "SHA1" in s or "SHA-1" in s:
        return "SHA 1"
    if "MD5" in s:
        return "MD5 OR ANYTHING ELSE"
    return "MD5 OR ANYTHING ELSE"

def _cert_size_severity(size_label):
    if size_label == "ANYTHING IN PQC":
        return "Low"
    if size_label in ("RSA 4096 & above", "RSA 3072", "RSA 2048", "secp384r1 & above", "secp256r1", "Ed25519", "Ed448"):
        return "Acceptable"
    if size_label == "RSA 1024":
        return "Medium"
    if size_label == "secp224r1":
        return "Medium"
    return "Critical"

def _sign_severity(sig_label):
    if sig_label == "ML-DSA, SLH DSA":
        return "Low"
    if sig_label in ("ECDSA + SHA256/384", "RSA + SHA256/384"):
        return "Acceptable"
    if sig_label == "ECDSA/RSA + SHA-1 or (S224)":
        return "Critical"
    if sig_label == "SHA 224":
        return "Medium"
    if sig_label == "SHA 1":
        return "High"
    if sig_label == "MD5 OR ANYTHING ELSE":
        return "Critical"
    return "Critical"

def _cert_key_severity(key_type):
    if key_type == "Dilithium/Falcon/SPHINCS (ML/SLH DSA)":
        return "Low"
    if key_type in ("EdDSA", "ECDSA", "RSA"):
        return "Acceptable"
    if key_type == "DSA":
        return "High"
    return "Critical"

def _cert_key_recommendation(key_type):
    if key_type == "Dilithium/Falcon/SPHINCS (ML/SLH DSA)":
        return "N.A."
    if key_type in ("EdDSA", "ECDSA", "RSA"):
        return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
    if key_type == "DSA":
        return "Strongly Advised to EdDSA/ECDSA with appropriate curve or ML DSA or SLH DSA for PQC environment"
    return "Strongly Advised to EdDSA/ECDSA with appropriate curve or ML DSA or SLH DSA for PQC environment"

def _cert_size_recommendation(size_label):
    if size_label == "ANYTHING IN PQC":
        return "N.A."
    if size_label in ("RSA 4096 & above", "RSA 3072", "secp384r1 & above", "secp256r1", "Ed25519", "Ed448"):
        return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
    if size_label == "RSA 2048":
        return "Recommended to RSA 3072 or ML DSA or SLH DSA for PQC environment"
    if size_label in ("RSA 1024", "secp224r1"):
        return "Strongly Advised to P-384 or ML DSA or SLH DSA for PQC environment"
    if size_label == "RSA 1024 & LESSER":
        return "Strongly Advised to RSA 3072 or ML DSA or SLH DSA for PQC environment"
    return "Strongly Advised to P-384 or ML DSA or SLH DSA for PQC environment"

def _signature_recommendation(sig_label):
    if sig_label == "ML-DSA, SLH DSA":
        return "N.A."
    if sig_label in ("ECDSA + SHA256/384", "RSA + SHA256/384"):
        return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
    if sig_label == "ECDSA/RSA + SHA-1 or (S224)":
        return "Strongly Advised to ECDSA + SHA256/384 or to ML DSA or SLH DSA for PQC environment"
    if sig_label == "SHA 224":
        return "Strongly Advised to ECDSA + SHA256/384 or to ML DSA or SLH DSA for PQC environment"
    if sig_label == "SHA 1":
        return "Strongly Advised to ECDSA + SHA256/384 or to ML DSA or SLH DSA for PQC environment"
    return "Strongly Advised to ECDSA + SHA256/384 or to ML DSA or SLH DSA for PQC environment"

def _norm(r, l, h): return 50.0 if h == l else round(max(0.0, min(100.0, (r - l) / (h - l) * 100)), 1)

def _rec(sev, act):
    if sev in ("Low", "Acceptable"): return "N.A."
    if sev == "Critical": return f"Strongly suggested to {act}"
    if sev == "High": return f"Recommended to {act}"
    if sev == "Medium": return f"Prefer to {act}"
    return f"N.A."

def _override_rec(param, finding, sev, rec):
    proto_recs = {
        "TLS 1.3": "N.A.",
        "V1.3": "N.A.",
        "TLS 1.2": "Recommended optional to v1.3",
        "V1.2": "Recommended optional to v1.3",
        "TLS 1.1": "Strongly Advised to v1.2 or above",
        "V1.1": "Strongly Advised to v1.2 or above",
        "TLS 1.0": "Strongly Advised to v1.2 or above",
        "V1.0": "Strongly Advised to v1.2 or above",
        "TLS 3.0": "Strongly Advised to TLSv1.2 or above",
        "V3.0": "Strongly Advised to TLSv1.2 or above",
        "TLS 2.0": "Strongly Advised to TLSv1.2 or above",
        "V2.0": "Strongly Advised to TLSv1.2 or above",
    }

    if str(param).startswith("Protocol - "):
        proto = str(param).split(" - ", 1)[1].strip()
        if proto in proto_recs:
            return proto_recs[proto]
        proto_norm = proto.replace("TLSV", "TLS ").replace("TLSv", "TLS ").upper()
        if proto_norm in proto_recs:
            return proto_recs[proto_norm]

    if param == "Downgrade Protection (TLS_FALLBACK_SCSV)":
        if finding == "Supported": return "N.A."
        if finding == "Not Supported": return "Recommended to enable TLS_FALLBACK_SCSV"

    if param == "SNI":
        if finding == "Supported": return "N.A."
        return "Enable TLS SNI support and Upgrade legacy TLS stacks that do not support RFC 6066."

    if param == "ALPN":
        if finding == "Supported": return "N.A."
        return "Enable ALPN on TLS endpoints, supports HTTP/2 at bare minimum and then enable HTTP/3."

    if param == "Extended Master Secret":
        if finding == "Supported": return "N.A."
        return "Prefer TLS 1.3 whether possible, otherwise Enable RFC 7627 Extended Master Secret support."

    if param == "Session Ticket":
        if finding == "Supported": return "Ensure frequent ticket key rotation and prefer TLS 1.3 resumption mechanisms."
        return "It is required for performance optimization."

    if param == "Renegotiation Info":
        if finding == "Supported": return "N.A."
        return "Enable RFC 5746 Secure Renegotiation."

    if param == "EC Point Formats":
        if finding == "Supported": return "N.A."
        return "Use uncompressed EC points only, Disable obsolete compressed point formats unless specifically required."

    if param == "Supported Versions":
        if finding == "Supported": return "N.A."
        return "This is important for TLS Version Negotiation, Downgrade Protection, TLS 1.3 Support, Modern Cryptography Adoption."

    if param == "Public Key Algorithm":
        algo = str(finding).upper()
        if any(x in algo for x in ("ML-DSA", "SLH-DSA", "FALCON", "DILITHIUM", "SPHINCS")):
            return "N.A."
        if "EDDSA" in algo:
            return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
        if "ECDSA" in algo:
            return "Recommended to ML DSA or SLH DSA for PQC environment"
        if "RSA" in algo:
            return "Recommended to ML DSA or SLH DSA for PQC environment"
        if "DSA" in algo:
            return "Strongly Advised to EdDSA/ECDSA with appropriate curve or ML DSA or SLH DSA for PQC environment"
        return "Strongly Advised to EdDSA/ECDSA with appropriate curve or ML DSA or SLH DSA for PQC environment"

    if param == "Public Key Size":
        size = str(finding)
        if "PQC" in size.upper():
            return "N.A."
        if "4096" in size:
            return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
        if "3072" in size:
            return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
        if "2048" in size:
            return "Recommended to RSA 3072 or ML DSA or SLH DSA for PQC environment"
        if "1024" in size and "LESSER" not in size:
            return "Strongly Advised to RSA 3072 or ML DSA or SLH DSA for PQC environment"
        if any(x in size.upper() for x in ("ED25519", "ED448")):
            return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
        if "256" in size and ("ECDSA" in str(finding).upper() or "EC" in str(finding).upper()):
            return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
        if "224" in size and ("ECDSA" in str(finding).upper() or "EC" in str(finding).upper()):
            return "Strongly Advised to P-384 or ML DSA or SLH DSA for PQC environment"
        return "Strongly Advised to RSA 3072 or ML DSA or SLH DSA for PQC environment"

    if param == "PQC Certificate DSA":
        if isinstance(finding, str) and any(x in finding.upper() for x in ("DILITHIUM", "ML-DSA", "MLDSA")):
            return "Implement ML-DSA for future certificate and signing infrastructure. Deploy hybrid certificates during migration."
        if isinstance(finding, str) and any(x in finding.upper() for x in ("FALCON", "FN-DSA", "FNDSA")):
            return "Consider FN-DSA where certificate size is critical. Evaluate implementation maturity."
        if isinstance(finding, str) and any(x in finding.upper() for x in ("SPHINCS", "SLH-DSA", "SLHDSA")):
            return "Maintain SLH-DSA as crypto-agility backup option. Use for high-assurance environments."
        if finding == "Classical":
            return "Enable PQC/hybrid key exchange support, Upgrade TLS libraries to versions supporting ML-KEM, Use hybrid TLS groups such as: X25519MLKEM768."
        return "N.A."

    if param == "PQC Key Encapsulation":
        if isinstance(finding, str) and finding.startswith("Pure PQC"):
            return "N.A."
        if isinstance(finding, str) and finding.startswith("Hybrid PQC"):
            return "N.A."
        if finding == "Not Supported (Classical)":
            return "Enable PQC/hybrid key exchange support, Upgrade TLS libraries to versions supporting ML-KEM, Use hybrid TLS groups such as: X25519MLKEM768."
        return "N.A."

    if param == "PQC Sinkhole":
        if finding == "Full PQC":
            return "N.A."
        if finding == "Classical / Partial":
            return "Because PQC authentication chain is incomplete, a degraded value was given. Recommended(OPT) to ML DSA or SLH DSA for PQC environment"
        return "N.A."

    if param == "Certificate Chain / Trust Store":
        if finding == "Trusted chain" or finding == "Trusted":
            return "N.A."
        if finding == "Self-Signed":
            return "Strongly Advised to have CA to chain"
        if finding == "Chain integrity lost" or finding == "Untrusted":
            return "Strongly Advised to check chain brokering or root stewardship"
        return "Strongly Advised to check chain brokering or root stewardship"

    if param == "Certificate Validity / Expiry":
        if isinstance(finding, str) and finding.startswith("Valid"):
            return "N.A."
        return "Strongly Advised to check certificate validity"

    if param == "Certificate OCSP Staple":
        if finding == "Supported and Valid":
            return "N.A."
        if finding == "Supported and Invalid":
            return "Strongly Advised to have OCSP Stapling validated from trusted authorities"
        return "Strongly Advised to have OCSP Stapling and be validated"

    if param == "Certificate Transparency":
        if finding == "3 SCT & more":
            return "N.A."
        if finding == "2 SCT":
            return "Recommended(OPT) to have atleast 3 SCT"
        if finding == "1 SCT":
            return "Strongly Advised to have atleast 3 SCT"
        if finding == "Absent":
            return "Strongly Advised to have SCTs, recommended atleast 3"
        return "Strongly Advised to have SCTs, recommended atleast 3"

    if param == "PQC Key Encapsulation":
        if finding == "Active":
            return "N.A."
        if finding == "Classical":
            return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"

    if param == "PQC Sinkhole":
        if finding == "Full PQC":
            return "N.A."
        if finding == "Classical / Partial":
            return "Recommended(OPT) to ML DSA or SLH DSA for PQC environment"

    return rec

class Section:
    def __init__(self, name, weight):
        self.name = name
        self.weight = weight
        self.rows = []

    def add(self, param, finding, std, sev, rec, lo, hi):
        rec = _override_rec(param, finding, sev, rec)
        self.rows.append({"param": param, "finding": finding, "std": std, "sev": sev, 
                          "score": SCORE.get(sev, 0), "rec": rec, "lo": lo, "hi": hi})

    def process(self):
        ts = sum(r["hi"] - r["lo"] for r in self.rows)
        for r in self.rows:
            r["wt"] = ((r["hi"] - r["lo"]) / ts * self.weight) if ts > 0 else 0
        raw = sum(r["score"] for r in self.rows)
        lo = sum(r["lo"] for r in self.rows)
        hi = sum(r["hi"] for r in self.rows)
        return raw, lo, hi, _norm(raw, lo, hi)

def _p_sev(p, s):
    if s:
        return {"TLS 1.3":"Low", "TLS 1.2":"Acceptable", "TLS 1.1":"High", "TLS 1.0":"Critical", "SSLv3":"Critical", "SSLv2":"Critical"}.get(p, "Acceptable")
    return {"TLS 1.3":"Critical", "TLS 1.2":"Medium", "TLS 1.1":"Low", "TLS 1.0":"Low", "SSLv3":"Low", "SSLv2":"Low"}.get(p, "Acceptable")

def _p_std(p): 
    return {"TLS 1.3":"NIST SP 800-52-r2", "TLS 1.2":"NIST SP 800-52-r2", "TLS 1.1":"RFC 8996", "TLS 1.0":"RFC 8996", "SSLv3":"RFC 7568", "SSLv2":"RFC 6176"}.get(p, "NIST SP 800-52r2")

def _pk_a(s, k):
    s, k = (s or "").upper(), k or 0
    if any(x in s for x in ("ML","SLH","FALCON","DILITHIUM","SPHINCS")): return "Low"
    if "ED25519" in s or "ED448" in s:
        return "Acceptable"
    if "ECDSA" in s or ("EC" in s and "RSA" not in s):
        return "Acceptable" if k >= 256 else "Medium" if k >= 224 else "Critical"
    if "RSA" in s:
        return "Acceptable" if k >= 2048 else "Medium" if k >= 1024 else "Critical"
    return "High"

def _exp(e, d):
    if e == "EXPIRED": return "Critical", -10
    try:
        dy = (datetime.fromisoformat(d.replace("Z","+00:00")) - datetime.now(timezone.utc)).days
        if dy < 0: return "Critical", -10
        if dy <= 30: return "Medium", -2
        if dy < 180: return "Acceptable", 5
        return "Low", 10
    except: return "Acceptable", 5

def _hsts(h):
    if not h or h == "not offered": return "High"
    if "preload" in h.lower(): return "Acceptable"
    m = re.search(r"max-age=(\d+)", h.lower())
    return "Acceptable" if (m and int(m.group(1)) >= 31536000) else "Medium"

def _cipher_key_exchange_recommendation(kx, bits=0):
    k = str(kx or "").upper()
    if any(x in k for x in ("MLKEM", "KYBER", "HYBRID", "X25519MLKEM")):
        return "ML-KEM (& ECDHE)", "CNSA 2.0 Standard", "N.A."
    if "ECDHE" in k:
        return "ECDHE", "NIST SP 800-52r2", "Recommended to ML-KEM (&ECDHE) for PQC"
    if "DHE" in k:
        if bits >= 2048:
            return "DHE >=2048", "NIST SP 800-131A", "Recommended to ECDHE for classic or ML-KEM for PQC"
        return "DHE <= 2048", "NIST SP 800-131A", "Strongly Advised to ECDHE for classic or ML-KEM for PQC"
    if "RSA" in k:
        return "RSA", "NIST SP 800-52r2", "Strongly Advised to ECDHE for classic or ML-KEM for PQC"
    if "STATIC" in k or ("DH" in k and "ECDHE" not in k and "DHE" not in k):
        return "Static DH", "NIST SP 800-131A", "Strongly Advised to ECDHE for classic or ML-KEM for PQC"
    if "ANON" in k:
        return "Anonymous DH", "NIST SP 800-131A", "Strongly Advised to ECDHE for classic or ML-KEM for PQC"
    return kx or "Unknown", "NIST SP 800-52r2", "Recommended to ML-KEM (&ECDHE) for PQC"

def _cipher_signature_recommendation(cipher_name, mac, bits=0):
    s = str(cipher_name or "").upper()
    m = str(mac or "").upper()
    if m == "SHA" and bits:
        return f"SHA-{bits}", "NIST SP 800-131A Rev. 2", "Strongly Advised to SHA-256(or SHA-3) or above"
    if m == "SHA1" or (m == "SHA" and bits == 1):
        return "SHA-1", "NIST SP 800-131A Rev. 2", "Strongly Advised to SHA-256(or SHA-3) or above"
    if "SHA512" in s or "SHA3" in s or (m == "SHA" and bits == 512):
        return "SHA-512 or SHA-3", "NIST SP 800-131A Rev. 2", "N.A."
    if "SHA384" in s or (m == "SHA" and bits == 384):
        return "SHA-384 or SHA -3", "NIST SP 800-131A Rev. 2", "N.A."
    if "SHA256" in s or (m == "SHA" and bits == 256):
        return "SHA-256", "NIST SP 800-131A Rev. 2", "N.A."
    if "SHA224" in s or (m == "SHA" and bits == 224):
        return "SHA-224", "NIST SP 800-131A Rev. 2", "Recommended optional to SHA256 or SHA 3"
    if "SHA1" in s or s.endswith("_SHA") or ("SHA" in s and "SHA256" not in s and "SHA384" not in s and "SHA512" not in s and "SHA3" not in s):
        return "SHA-1", "NIST SP 800-131A Rev. 2", "Strongly Advised to SHA-256(or SHA-3) or above"
    if "MD5" in s and "SHA1" in s:
        return "MD5 + SHA1", "NIST SP 800-131A Rev. 2", "Strongly Advised to SHA-256(or SHA-3) or above"
    if "MD5" in s:
        return "MD5", "NIST SP 800-131A Rev. 2", "Strongly Advised to SHA-256(or SHA-3) or above"
    if m == "SHA" and bits:
        return f"SHA-{bits}", "NIST SP 800-131A Rev. 2", "Strongly Advised to SHA-256(or SHA-3) or above"
    return "SHA-256", "NIST SP 800-131A Rev. 2", "N.A."

def _cipher_bulk_recommendation(bulk, mac):
    b = str(bulk or "").upper()
    if "CHACHA20" in b:
        return "ChaCha20-Poly1305", "NIST FIPS 197", "N.A."
    if "AES-256-GCM" in b:
        return "AES-256-GCM", "RFC 8439", "N.A."
    if "AES-128-GCM" in b:
        return "AES-128-GCM", "RFC 8439", "Recommended(OPT) to have ChaCha20 or AES 256"
    if "AES CCM" in b or "AES-CCM" in b:
        return "AES CCM", "NIST SP 800-38D", "Recommended to have GCM or Poly1305"
    if "3DES" in b:
        return "3DES CBC", "NIST SP 800-131A rev 2", "Strongly Advised to have ChaCha20 or AES 128 atleast"
    if "RC4" in b:
        return "RC4", "RFC 7465 / RFC 5469", "Strongly Advised to have ChaCha20 or AES 128 atleast"
    if "NULL" in b or not b:
        return "NULL", "RFC 5246", "Strongly Advised to have ChaCha20 or AES 128 atleast"
    if "CBC" in b:
        return "AES CBC + HMAC", "NIST SP 800-131A rev 2/RFC 5246", "Strongly Advised to have ChaCha20 or AES 128 atleast"
    return b or "Unknown", "RFC 8439", "Recommended(OPT) to have ChaCha20 or AES 256"

def _normalized_negotiated_group_name(group):
    if not group:
        return ""
    g = str(group).strip()
    if not g:
        return ""
    if re.search(r"x25519.*mlkem|mlkem.*x25519", g, re.I):
        return "(X25519)MLKEM768"
    if re.search(r"\bx25519\b", g, re.I):
        return "X25519"
    if re.search(r"\bsecp256r1\b", g, re.I):
        return "secp256r1"
    if re.search(r"\bsecp384r1\b", g, re.I):
        return "secp384r1"
    if re.search(r"\bsecp521r1\b", g, re.I):
        return "secp521r1"
    if re.search(r"\bsecp224r1\b", g, re.I):
        return "secp224r1"
    if re.search(r"\bsecp192r1\b", g, re.I):
        return "secp192r1"
    m = re.search(r"\bffdhe(\d+)\b", g, re.I)
    if m:
        size = int(m.group(1))
        if size >= 4096:
            return "ffdhe4096"
        if size >= 3072:
            return "ffdhe3072"
        if size == 2048:
            return "ffdhe2048"
        return "ffdhe2048 & LESSER"
    if re.search(r"kb|kyber|mlkem|hybrid", g, re.I):
        return "(X25519)MLKEM768"
    return g


def _negotiated_group_info(group):
    g = _normalized_negotiated_group_name(group)
    if g == "(X25519)MLKEM768":
        return g, "CNSA 2.0 Standard", "Low", "N.A."
    if g == "X25519":
        return g, "NIST SP 800-186", "Acceptable", "Recommended to MLKEM786(or hybrid) for PQC"
    if g == "secp256r1":
        return g, "NIST SP 800-186", "Acceptable", "Recommended to MLKEM786(or hybrid) for PQC"
    if g == "secp384r1":
        return g, "NIST SP 800-186", "Acceptable", "Recommended to MLKEM786(or hybrid) for PQC"
    if g == "secp521r1":
        return g, "NIST SP 800-186", "Acceptable", "Recommended to MLKEM786(or hybrid) for PQC"
    if g == "ffdhe4096":
        return g, "RFC 7919", "Acceptable", "Recommended to MLKEM786(or hybrid) for PQC"
    if g == "ffdhe3072":
        return g, "RFC 7919", "Acceptable", "Recommended to MLKEM786(or hybrid) for PQC"
    if g == "ffdhe2048":
        return g, "RFC 7919", "Medium", "Strongly Advised to ffdhe4096 or MLKEM786(or hybrid) for PQC"
    if g == "ffdhe2048 & LESSER":
        return g, "RFC 7919", "Critical", "Strongly Advised to ffdhe4096 or MLKEM786(or hybrid) for PQC"
    if g == "secp224r1":
        return g, "RFC 7919", "High", "Strongly Advised to X25519 or MLKEM786(or hybrid) for PQC"
    if g == "secp192r1":
        return g, "RFC 7919", "Critical", "Strongly Advised to X25519 or MLKEM786(or hybrid) for PQC"
    return g, "NIST SP 800-186", "Medium", "Review negotiated group support and consider MLKEM or X25519 based hybrid transitions."


def _collect_negotiated_groups(data):
    groups = []
    seen = set()

    def push_name(name):
        if not name:
            return
        norm = _normalized_negotiated_group_name(name)
        if norm and norm not in seen:
            seen.add(norm)
            groups.append(norm)

    # extract explicit group names from cipher key_exchange entries
    for cs in data.get("cipher_suites", []):
        if not isinstance(cs, dict):
            continue
        kx = str(cs.get("key_exchange", "") or "")
        # also accept explicit curve lists inside each cipher suite
        if isinstance(cs.get("ecdhe_curves"), list):
            for c in cs.get("ecdhe_curves"):
                push_name(c)
        # some scanners put supported groups in different keys
        if isinstance(cs.get("supported_groups"), list):
            for c in cs.get("supported_groups"):
                push_name(c)
        # find explicit curve or ffdhe or pqc tokens inside the key exchange string
        m = re.search(r"(X25519|SECP256R1|SECP384R1|SECP521R1|SECP224R1|SECP192R1|FFDHE\d+|KYBER|MLKEM)", kx, re.I)
        if m:
            push_name(m.group(0))
        else:
            # sometimes key_exchange contains forms like 'ECDHE-RSA' with curve in auth or elsewhere
            # try parsing common curve tokens separated by non-alphanum
            parts = re.split(r"[^A-Za-z0-9()_-]+", kx)
            for p in parts:
                if re.search(r"^(X25519|SECP\d+R\d+|FFDHE\d+|KYBER|MLKEM)$", p, re.I):
                    push_name(p)

    # include PQC negotiated group if available
    pq = data.get("pqc_active_probe")
    if isinstance(pq, dict):
        ng = pq.get("negotiated_group", "") or ""
        if ng:
            # pq negotiated group may include multiple tokens
            # split on commas or spaces and try to normalize each
            for tok in re.split(r"[,;\s]+", str(ng)):
                push_name(tok)

    # include explicit ECDHE curve list if present in CBOM (common keys: 'ecdhe_curves' under top-level, 'tls', or 'pfs')
    ecdhe_list = data.get("ecdhe_curves") or (data.get("tls", {}) or {}).get("ecdhe_curves") or (data.get("pfs", {}) or {}).get("ecdhe_curves")
    if isinstance(ecdhe_list, list):
        for c in ecdhe_list:
            push_name(c)

    return groups


def _build_rs(wb, data):
    ws = wb.create_sheet("Posture Score")
    ws.sheet_view.showGridLines = False
    
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    
    sm = data.get('scan_metadata')
    sm = sm if isinstance(sm, dict) else {}
    t_val = sm.get('target') or data.get('target', {})
    
    if isinstance(t_val, dict):
        host_name = t_val.get('hostname', str(t_val.get('host', 'Unknown')))
        ip_addr = t_val.get('ip', '')
    else:
        host_name = str(t_val) if t_val else 'Unknown'
        ip_addr = ''
        
    if not ip_addr:
        ip_addr = sm.get('ip', sm.get('server_ip', data.get('ip', '')))
        
    display_title = f"{host_name} ({ip_addr})" if ip_addr else host_name
    
    c.value = "AUDIT REPORT"
    c.font = _title_font(C["TITLE_FG"], b=False, s=14, u="single"); c.fill = _f(C["TITLE_BG"])
    c.alignment = CTR; c.border = BRD; ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Hostname: {host_name}"
    ws["A2"].font = _font(C["COL_FG"], s=11)
    ws["A2"].alignment = WRP; ws["A2"].border = BRD

    ws.merge_cells("A3:G3")
    ws["A3"].value = f"IP: {ip_addr}"
    ws["A3"].font = _font(C["COL_FG"], s=11)
    ws["A3"].alignment = WRP; ws["A3"].border = BRD

    ws.append([""])
    cols = ["Parameter", "Finding / Observed Value", "Standard Mapping", "Severity", "Score", "Contribution to posture score", "Recommendation"]
    ws.append(cols)
    for i, cell in enumerate(ws[ws.max_row], 1):
        cell.font = _font(C["COL_FG"], b=True); cell.fill = _f(C["COL_BG"]); cell.alignment = CTR; cell.border = BRD
        ws.column_dimensions[get_column_letter(i)].width = [32, 34, 30, 13, 8, 12, 100][i-1]
    ws.row_dimensions[ws.max_row].height = 20; ws.freeze_panes = "A4"

    res = []
    def _flush(s):
        ws.append([""]); r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c = ws.cell(r, 1, f"  {s.name}  (Weight: {s.weight}%)")
        c.font = _font(C["SEC_FG"], b=True); c.fill = _f(C["SEC_BG"]); c.alignment = WRP; c.border = BRD
        ws.row_dimensions[r].height = 18
        
        rw, lo, hi, nm = s.process()
        for v in s.rows:
            ws.append([v["param"], v["finding"], v["std"], v["sev"], v["score"], v["wt"]/100, v["rec"]])
            mr = ws.max_row
            for ci, cell in enumerate(ws[mr]):
                cell.border = BRD; cell.alignment = WRP; cell.font = _font()
            sev_bg, sev_fg = SEV_STYLE.get(v["sev"], ("00000000", "00000000"))
            ws[mr][3].fill = _f(sev_bg); ws[mr][3].font = _font(sev_fg, b=True); ws[mr][3].alignment = CTR
            ws[mr][4].alignment = RGT; ws[mr][4].font = _font();
            ws[mr][4].number_format = '0.0;-0.0;0.0'
            contribution = round(v["score"] * v["wt"] / 10.0, 2)
            ws[mr][5].value = contribution / 100.0
            ws[mr][5].number_format = '+0.00%;-0.00%;+0.00%'
            ws[mr][5].alignment = RGT
            ws[mr][5].font = _font()
        res.append((s.name, rw, lo, hi, nm, s.weight))

    s = Section("Protocol (Capability space)", 23)
    for p in data.get("protocols", []):
        if not isinstance(p, dict): continue
        pr, sup = p.get("protocol", ""), p.get("supported", False)
        sv = _p_sev(pr, sup)
        
        if pr in ("SSLv2", "SSLv3", "TLS 1.0"): b = (-10, 10)
        elif pr == "TLS 1.1": b = (-6, 10)
        elif pr == "TLS 1.2": b = (-2, 5)
        elif pr == "TLS 1.3": b = (-10, 10)
        else: b = (-10, 10)
        
        act = f"disable {pr} immediately." if sup and pr in ("SSLv2", "SSLv3", "TLS 1.0", "TLS 1.1") else f"enable {pr}."
        s.add(f"Protocol - {pr} (Capability space)", "Enabled" if sup else "Disabled", _p_std(pr), sv, _rec(sv, act), *b)
        
    vu = data.get("vulnerabilities")
    vu = vu if isinstance(vu, dict) else {}
    fb = vu.get("Fallback_SCSV")
    fb = fb if isinstance(fb, dict) else {}
    
    fb_sv = "Low" if fb.get("is_supported") else ("High" if fb.get("legacy_protocol_present") else "Acceptable")
    s.add("Downgrade Protection (TLS_FALLBACK_SCSV)", "Supported" if fb.get("is_supported") else "Not Supported", "RFC 7507", fb_sv, _rec(fb_sv, "enable TLS_FALLBACK_SCSV to prevent protocol downgrade attacks."), -2, 5)
    _flush(s)

    s = Section("Certificate", 13)
    certs = [c for c in data.get("certificates", []) if isinstance(c, dict)]
    if not certs:
        certs = [{}]

    for idx, cert in enumerate(certs, 1):
        sig = cert.get("Signature Algorithm", "")
        raw_ksz = cert.get("Public Key Size", 0) or 0
        ksz = _normalize_cert_key_size(cert.get("cbom_authentication_layer", {}).get("algorithm_name", "") or sig, raw_ksz)
        cb_algo = ""
        if isinstance(cert.get("cbom_authentication_layer"), dict):
            cb_algo = cert["cbom_authentication_layer"].get("algorithm_name", "") or ""
        cert_algo_ref = cb_algo if cb_algo else sig
        algo_display = _cert_algo_display(sig, cb_algo, ksz)
        key_type = _cert_key_type(sig, cert_algo_ref)
        size_label = _cert_size_label(cert_algo_ref, ksz)
        signature_label = _signature_label(sig)

        s.add(f"Certificate {idx} - Public Key Cert", algo_display, "FIPS 204 / NIST SP 800-131A rev 2", _cert_key_severity(key_type), _cert_key_recommendation(key_type), -10, 10)
        s.add(f"Certificate {idx} - Public Keysize Cert", size_label, "NIST SP 800-131A rev 2", _cert_size_severity(size_label), _cert_size_recommendation(size_label), -10, 10)
        s.add(f"Certificate {idx} - Signature", signature_label, "NIST SP 800-131A rev 2", _sign_severity(signature_label), _signature_recommendation(signature_label), -10, 10)

        eku_raw = cert.get("Extended Key Usage", [])
        eku = [k.lower() for k in eku_raw] if isinstance(eku_raw, list) else []
        eku_sv = "Low" if any("server" in k for k in eku) else "High"
        s.add(f"Certificate {idx} - Extended Key Usage", _j(eku) or "None", "RFC 5280", eku_sv, _rec(eku_sv, "ensure serverAuth is present in EKU extensions."), -6, 10)

        es, n_a = cert.get("Expiration Status", ""), cert.get("Valid Not After", "")
        x_sev, _ = _exp(es, n_a)
        valid_label = "Valid" if isinstance(es, str) and "VALID" in es.upper() else "Invalid"
        date_display = ""
        if n_a:
            try:
                dt = datetime.fromisoformat(str(n_a).replace("Z", "+00:00"))
                date_display = dt.strftime("%d-%b-%Y")
            except Exception:
                date_display = str(n_a)[:10]
        rec_text = "renew certificate immediately to avoid service outage."
        if x_sev == "Medium":
            rec_text = "Please renew certificate soon as it expires within 30 days."
        s.add(f"Certificate {idx} - Validity / Expiry", f"{valid_label} ({date_display})", "CAB Forum Baseline Requirements", x_sev, _rec(x_sev, rec_text), -10, 10)

        ocsp_urls = cert.get("OCSP URLs") if isinstance(cert.get("OCSP URLs"), list) else []
        ocsp_finding = "Supported and Valid" if ocsp_urls else "Not Supported"
        ocsp_sv = "Low" if ocsp_urls else "High"
        s.add(f"Certificate {idx} - OCSP Staple", ocsp_finding, "NIST SP 800-52r2", ocsp_sv, _rec(ocsp_sv, "have OCSP Stapling validated from trusted authorities."), -6, 5)

        ct = cert.get("Certificate Transparency", "")
        ct_lower = str(ct).lower()
        if "3" in ct_lower or "sct present" in ct_lower:
            ct_label = "3 SCT & more"
        elif "2" in ct_lower:
            ct_label = "2 SCT"
        elif "1" in ct_lower:
            ct_label = "1 SCT"
        else:
            ct_label = "Absent"
        ct_sv = "Low" if ct_label == "3 SCT & more" else "Medium" if ct_label == "2 SCT" else "High"
        s.add(f"Certificate {idx} - Certificate Transparency", ct_label, "Google Policy", ct_sv, _rec(ct_sv, "enable Certificate Transparency logging."), -6, 5)

    tr = data.get("trust_stores")
    tr = tr if isinstance(tr, dict) else {}
    trusted = any(v.get("is_trusted") for v in tr.values() if isinstance(v, dict))
    chain_label = "Trusted chain" if trusted else ("Self-Signed" if len(certs) == 1 else "Chain integrity lost")
    chain_sv = "Low" if trusted else "High"
    s.add("Certificate Chain / Trust Store", chain_label, "PKI BMP, RFC 5280", chain_sv, _rec(chain_sv, "fix incomplete certificate chain or untrusted root CA."), -6, 10)
    _flush(s)

    s = Section("Extensions", 3)
    tx = data.get("tls_extensions")
    tx = tx if isinstance(tx, list) else []
    ext_data = [
        ("SNI", "server name", "NIST SP 800-52 Rev.2/RFC 6066", "implement SNI for multi-tenant hosting to ensure correct certificate routing."),
        ("ALPN", "application layer", "NIST SP 800-52 Rev.2/RFC 7301", "enable ALPN to negotiate application protocols securely."),
        ("Extended Master Secret", "extended master", "NIST SP 800-52 Rev.2/RFC 7627", "enable EMS to prevent MITM session hash vulnerabilities."),
        ("Session Ticket", "session ticket", "RFC 5077", "ensure keys rotate frequently if session tickets are enabled."),
        ("Renegotiation Info", "renegotiation", "CVE-2009-3555/ RFC 5746", "enable secure renegotiation (RFC 5746) to prevent injection attacks."),
        ("EC Point Formats", "EC point", "RFC 4492", "support standard ECC point formats for client compatibility."),
        ("Supported Versions", "supported version", "RFC 8446", "ensure TLS 1.3 supported versions match intended protocols.")
    ]
    for en, kw, std, act in ext_data:
        pr = any(isinstance(e, str) and kw.lower() in e.lower() for e in tx)
        sv = "Acceptable" if pr else "Medium"
        s.add(en, "Supported" if pr else "Not Supported", std, sv, _rec(sv, act), -2, 5)
    _flush(s)

    s = Section("HTTP Security", 2)
    ht = data.get("http_response")
    ht = ht if isinstance(ht, dict) else {}
    
    hs = ht.get("HSTS","not offered")
    hs_sv = _hsts(str(hs) if hs else "")
    s.add("HSTS", hs, "RFC 6797", hs_sv, _rec(hs_sv, "enforce HSTS header with max-age >= 31536000."), -6, 5)
    
    sc, hc = ht.get("Secure Cookie",""), ht.get("HTTP-only Cookie","")
    sc_str = str(sc).lower() if sc else ""
    hc_str = str(hc).lower() if hc else ""
    
    if "all" in sc_str:
        cookie_sv = "Acceptable"
        cookie_act = "N.A."
    else:
        cookie_sv = "High"
        cookie_act = _rec(cookie_sv, "set Secure and HttpOnly flags on application cookies.")
    s.add("Cookie Security", f"Secure: {sc} | HttpOnly: {hc}", "RFC 6265", cookie_sv, cookie_act, -6, 5)
    
    bn = ht.get("Server Banner","")
    bn_vuln = re.search(r"[\d.]+", str(bn)) if bn else False
    bn_sv = "Medium" if bn_vuln else "Acceptable"
    s.add("Server Banner", bn or "None", "NIST SP 800-44 / OWASP", bn_sv, _rec(bn_sv, "suppress verbose server version details in headers."), -2, 5)
    _flush(s)

    s = Section("Vulnerabilities", 18)
    vuln_meta = {
        "Heartbleed": ("Critical", "CVE-2014-0160", "Upgrade OpenSSL immediately, Revoke and reissue certificates, Rotate private keys, Invalidate active sessions, Force password resets if compromise is suspected.", (-10,10)),
        "ROBOT": ("Critical", "CVE-2017-13099", "Disable RSA Key Exchange cipher suites, Use ECDHE/X25519 key exchange.", (-10,10)),
        "CCS_Injection": ("Critical", "CVE-2014-0224", "Update OpenSSL/TLS libraries immediately, Patch vulnerable systems.", (-10,10)),
        "POODLE_SSL": ("Critical", "CVE-2014-3566", "Disable SSL 3.0 completely, Enable TLS 1.2 and TLS 1.3 only, Remove legacy CBC-only configurations.", (-10,10)),
        "SWEET32": ("High", "CVE-2016-2183", "Disable all 64-bit block cipher, remove 3DES, use AES-GCM, prefer ChaCha20-Poly1305, enable TLS 1.3.", (-6,10)),
        "BEAST": ("High", "CVE-2011-3389", "Disable TLS 1.0, SSL 3.0, legacy CBC ciphers, prefer TLS 1.2, 1.3 and Use ChaCha20-Poly1305 or AES-GCM.", (-6,10)),
        "CRIME_TLS": ("High", "CVE-2012-4929", "disable TLS level compression to prevent information leakage.", (-6,10)),
        "DROWN": ("Critical", "CVE-2016-0800", "disable SSLv2 immediately to secure RSA keys.", (-10,10)),
        "FREAK": ("High", "CVE-2015-0204", "disable EXPORT grade RSA ciphers.", (-6,10)),
        "RC4_Flaw": ("High", "CVE-2013-2566", "disable all RC4 cipher suites completely.", (-6,10)),
        "NULL_Cipher": ("Critical", "CVE-2002-20001", "disable NULL cipher suites that provide no encryption.", (-10,10)),
        "Secure_Renegotiation": ("High", "RFC 5746 / CVE-2009-3555", "enable RFC 5746 secure renegotiation to prevent MiTM injection.", (-6,10))
    }
    for vn, (sv, std, act, lb) in vuln_meta.items():
        if vn not in vu: continue
        vd = vu[vn]
        if not isinstance(vd, dict): continue
        v_s = "Low" if vn=="Secure_Renegotiation" and vd.get("is_supported") else "High" if vn=="Secure_Renegotiation" else sv if vd.get("is_vulnerable") else "Low"
        s.add(vn.replace("_"," "), "Vulnerable" if vd.get("is_vulnerable") else "Safe", std, v_s, _rec(v_s, act), *lb)
    _flush(s)

    s = Section("Cipher Categories", 2)
    cc_list = data.get("cipher_categories", [])
    cm = {c.get("category",""):c.get("supported",False) for c in cc_list if isinstance(c, dict)}
    c_meta = {
        "NULL Ciphers": ("Critical", "RFC 8996", "remove all NULL ciphers as they offer zero encryption.", (-10,10)),
        "Anonymous NULL Ciphers": ("Critical", "RFC 8996", "remove unauthenticated ciphers to prevent MITM attacks.", (-10,10)),
        "Export Ciphers": ("Critical", "CVE-2015-0204", "remove legacy EXPORT ciphers to prevent FREAK attacks.", (-10,10)),
        "LOW Ciphers (64-bit)": ("Critical", "RFC 4346", "remove weak 64-bit encryption ciphers.", (-10,10)),
        "RC4": ("Critical", "RFC 7465", "disable RC4 algorithms entirely due to statistical biases.", (-10,10)),
        "3DES / IDEA": ("High", "NIST SP 800-131A", "migrate away from obsolete block ciphers to AES or ChaCha20.", (-6,10)),
        "Obsolete CBC Ciphers": ("Medium", "NIST SP 800-52r2", "phase out CBC suites in favor of AEAD (GCM/Poly1305).", (-6,5)),
        "Strong Encryption (AEAD)": ("Low", "RFC 5116", "prioritize AEAD ciphers for maximum integrity and performance.", (-6,10))
    }
    for cn, (bs, std, act, bnd) in c_meta.items():
        pr = cm.get(cn, False)
        sv = "Low" if cn=="Strong Encryption (AEAD)" and pr else "High" if cn=="Strong Encryption (AEAD)" else bs if pr else "Low"
        s.add(cn, "Present" if pr else "Absent", std, sv, _rec(sv, act), *bnd)
    _flush(s)

    cs_sec = Section("Cipher Suites", 23)
    cs_list = data.get("cipher_suites", [])
    for cs in cs_list:
        if not isinstance(cs, dict): continue
        nm = cs.get("cipher_name","").upper()
        sv = "Critical" if any(x in nm for x in ("NULL","ANON","EXPORT","RC4")) else "High" if any(x in nm for x in ("3DES","IDEA")) else "Medium" if "CBC" in nm and not cs.get("aead") else "Low" if cs.get("aead") and cs.get("forward_secrecy") else "Acceptable" if cs.get("aead") else "Medium" if not cs.get("forward_secrecy") else "Acceptable"
        cs_sec.add(cs.get("cipher_name",""), "Supported", "RFC 8446 / RFC 5246", sv, "Review in Cipher Suites Tab.", -10, 10)
    
    rw, lo, hi, nm = cs_sec.process()
    res.append((cs_sec.name, rw, lo, hi, nm, cs_sec.weight))
    
    ws.append([""]); r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws.cell(r, 1, f"  {cs_sec.name}  (Weight: {cs_sec.weight}%)")
    c.font = _font(C["SEC_FG"], b=True); c.fill = _f(C["SEC_BG"]); c.alignment = WRP; c.border = BRD
    ws.row_dimensions[r].height = 18
    ws.append(["View detailed cipher breakdown", "Multiple Suites Evaluated", "RFC 8446 / RFC 5246", "Acceptable", SCORE.get("Acceptable", 5), cs_sec.weight/100, "Refer to the separate 'Cipher Suites' sheet for full breakdown."])
    for ci, cell in enumerate(ws[ws.max_row]):
        cell.border = BRD; cell.alignment = CTR; cell.font = _font()
    ws[ws.max_row][0].alignment = WRP; ws[ws.max_row][6].alignment = WRP
    ws[ws.max_row][5].number_format = '0.0%'

    s = Section("Negotiated Groups (Capability space)", 0)
    groups = _collect_negotiated_groups(data)
    if groups:
        for idx, group in enumerate(groups):
            param = "Negotiated Groups (Curves, field DH, PQC)" if idx == 0 else ""
            label, std, sev, rec = _negotiated_group_info(group)
            s.add(param, label, std, sev, rec, -2, 5)
    else:
        s.add("Negotiated Groups (Curves, field DH, PQC)", "None found", "N/A", "N/A", "N.A.", -2, 5)
    _flush(s)

    s = Section("PQC", 16)
    pq = data.get("pqc_active_probe")
    pq = pq if isinstance(pq, dict) else {}
    
    c0 = certs[0] if certs else {}
    cb = c0.get("cbom_authentication_layer")
    cb = cb if isinstance(cb, dict) else {}
    
    kx = pq.get("pqc_negotiated", False)
    cp = cb.get("is_pqc", False)
    
    sig_alg = c0.get("Signature Algorithm", "").upper()
    ds_pqc = any(x in sig_alg for x in ("ML-DSA", "DILITHIUM", "SLH-DSA", "SPHINCS", "FALCON"))
    ds_sv = "Low" if ds_pqc else "Medium"
    
    kx_sv = "Low" if kx else "Medium"
    neg_group = pq.get("negotiated_group", "") or ""
    neg_group_uc = str(neg_group).upper()
    if kx:
        if "MLKEM" in neg_group_uc or "KYBER" in neg_group_uc:
            kx_finding = f"Hybrid PQC supported: {neg_group}"
        else:
            kx_finding = f"Hybrid PQC supported: {neg_group}" if neg_group_uc else "Hybrid PQC supported"
    else:
        kx_finding = "Hybrid PQC not supported"
    s.add("PQC Key Encapsulation", kx_finding, "NIST FIPS 203", kx_sv, _rec(kx_sv, "plan infrastructure migration to ML-KEM (FIPS 203)."), -2, 10)

    if ds_pqc:
        if "DILITHIUM" in sig_alg:
            ds_finding = "Pure PQC: DSA(Dilithium) — ML DSA"
        elif "FALCON" in sig_alg or "FN-DSA" in sig_alg or "FNDSA" in sig_alg:
            ds_finding = "Pure PQC: DSA(Falcon) — FL DSA"
        elif "SPHINCS" in sig_alg or "SLH" in sig_alg:
            ds_finding = "Pure PQC: DSA(SPHINCS+) — SH DSA"
        else:
            ds_finding = "Pure PQC: DSA — ML DSA"
    else:
        ds_finding = "Classical"
    s.add("PQC Certificate DSA", ds_finding, "NIST FIPS 204, FIPS 205, FIPS 206", ds_sv, _rec(ds_sv, "plan certificate migration to ML-DSA or SLH-DSA."), -2, 10)
    
    sink_sv = "Low" if kx and cp else "Medium"
    sink_finding = "Full PQC" if kx and cp else "Classical / Partial"
    s.add("PQC Sinkhole", sink_finding, "NIST SP 800-52", sink_sv, _rec(sink_sv, "enable full PQC authentication chain."), -2, 10)
    
    tel = pq.get("negotiation_telemetry")
    tel = tel if isinstance(tel, dict) else {}
    
    is_downgraded = tel.get("downgrade_detected", False)
    pqc_disabled = not kx and not ds_pqc
    
    if pqc_disabled or is_downgraded:
        dg_sv = "Medium"
        dg_finding = "PQC Disabled / Forced Classical" if pqc_disabled else "Downgrade Detected"
    else:
        dg_sv = "Low"
        dg_finding = "Well Supported"
        
    s.add("PQC Downgrade", dg_finding, "RFC 8446", dg_sv, _rec(dg_sv, "investigate potential MITM protocol downgrade attack."), -2, 10)
    _flush(s)

    ws.append([""]); r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    hc = ws.cell(r, 1, "  Section Summary & Posture Score")
    hc.font = _font("FF000000", b=True, s=11); hc.fill = _f(C["TITLE_BG"]); hc.border = BRD
    ws.append(["Section", "Weight", "Raw Score", "Min Possible", "Max Possible", "Normalised", ""]); hr = ws.max_row
    for c in ws[hr]: c.font = _font(C["COL_FG"], b=True); c.fill = _f(C["COL_BG"]); c.alignment = CTR; c.border = BRD
    ws.merge_cells(start_row=hr, start_column=6, end_row=hr, end_column=7)
    
    wt_sum = sum(nm * (wt/100) for _, _, _, _, nm, wt in res)
    for (lb, rw, lo, hi, nm, wt) in res:
        ws.append([lb, wt/100, rw, lo, hi, nm, ""])
        sr = ws.max_row; ws.merge_cells(start_row=sr, start_column=6, end_row=sr, end_column=7)
        ws[sr][1].number_format = '0.0%'
        bg = "FFC6EFCE" if nm>=70 else "FFFFC7CE" if nm<40 else "FFFFEB9C"
        for c in ws[sr]: c.border = BRD; c.alignment = CTR; c.font = _font(); c.fill = _f(bg)
        ws[sr][0].alignment = WRP
    ws.append(["Posture score:", 1.0, "", "", "", round(wt_sum, 2), ""])
    fr = ws.max_row; ws.merge_cells(start_row=fr, start_column=6, end_row=fr, end_column=7)
    ws[fr][1].number_format = '0.0%'
    for c in ws[fr]: c.font = _font("FF000000", b=True, s=12); c.fill = _f(C["TITLE_BG"]); c.alignment = CTR; c.border = BRD
    ws[fr][0].alignment = WRP

def _build_cs(wb, data):
    ws = wb.create_sheet("Cipher Suites (Capability)")
    ws.sheet_view.showGridLines = False
    cl = ["#", "Cipher Suite", "Protocol", "Key Exchange", "Authentication", "Bulk Encryption", "Handshake Hash", "Perfect Forward Secrecy", "Bits", "Severity", "Score", "Recommendation"]
    ws.merge_cells("A1:L1")
    t = ws["A1"]; t.value = "Cipher Suite Assessment (Capability space)"; t.font = _font(C["TITLE_FG"], b=True, s=12); t.fill = _f(C["TITLE_BG"]); t.alignment = CTR; t.border = BRD
    ws.append(cl); hr = ws.max_row
    widths = [4, 46, 10, 15, 15, 18, 30, 25, 8, 12, 8, 120]
    for i, c in enumerate(ws[hr], 1): 
        c.font = _font(C["COL_FG"], b=True); c.fill = _f(C["COL_BG"]); c.alignment = CTR; c.border = BRD
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.freeze_panes = "A3"
    
    cs_list = data.get("cipher_suites", [])
    for i, cs in enumerate(cs_list, 1):
        if not isinstance(cs, dict): continue
        nm = cs.get("cipher_name","" ).upper()
        kx = str(cs.get("key_exchange","")) or ""
        auth = str(cs.get("authentication","")) or ""
        pfs = cs.get("forward_secrecy")
        sec_bits = cs.get("security_bits", 0) or 0
        if any(x in nm for x in ("NULL","ANON","EXPORT","RC4")):
            sv = "Critical"
        elif "RSA" in kx and "ECDHE" not in kx and "DHE" not in kx:
            sv = "High"
        elif "STATIC" in kx or ("DH" in kx and "ECDHE" not in kx and "DHE" not in kx):
            sv = "High"
        elif "DHE" in kx and "ECDHE" not in kx:
            sv = "Acceptable" if any(d in nm for d in ("DHE_RSA","DHE_DSS","DHE_ANON")) and sec_bits >= 2048 else "High"
        elif "ECDHE" in kx and "MLKEM" not in kx and "DHE" not in kx:
            sv = "Low"
        elif any(x in kx for x in ("MLKEM", "KYBER", "HYBRID")):
            sv = "Low"
        else:
            sv = "Acceptable" if cs.get("aead") else "Medium"
        key_label, key_std, key_rec = _cipher_key_exchange_recommendation(kx, sec_bits)
        sig_label, sig_std, sig_rec = _cipher_signature_recommendation(nm, cs.get("mac",""), sec_bits)
        bulk_label, bulk_std, bulk_rec = _cipher_bulk_recommendation(cs.get("bulk_encryption",""), cs.get("mac",""))
        recommendation = f"{key_label} ({key_std}): {key_rec}, {sig_label} ({sig_std}): {sig_rec}, {bulk_label} ({bulk_std}): {bulk_rec}"
        pfs_label = "Yes (Ephemeral)" if pfs else "Not Supported"
        ws.append([i, cs.get("cipher_name",""), cs.get("protocol",""), cs.get("key_exchange",""), cs.get("authentication",""), cs.get("bulk_encryption",""), sig_label, pfs_label, cs.get("security_bits",""), sv, SCORE.get(sv, 0), recommendation])
        r = ws.max_row
        for c in ws[r]: c.border = BRD; c.alignment = WRP; c.font = _font()
        sb, sf = SEV_STYLE.get(sv, ("00000000","00000000"))
        ws[r][9].fill = _f(sb); ws[r][9].font = _font(sf, b=True); ws[r][9].alignment = CTR
        ws[r][10].alignment = CTR; ws[r][10].font = _font(b=True)
        ws[r][7].alignment = CTR; ws[r][7].font = _font("FF2E7D32" if pfs else "FFC62828", b=True)

def generate_risk_report(json_path: str, output_path: str) -> bool:
    if not os.path.exists(json_path):
        print(f"[-] ERROR: File not found -> {json_path}")
        return False
    try:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
        _build_rs(wb, data)
        _build_cs(wb, data)
        wb.save(output_path)
        print(f"[+] Successfully generated: {output_path}")
        return True
    except Exception as e:
        print(f"[-] CRASH: Excel generation failed -> {e}")
        return False

if __name__ == "__main__":
    j = sys.argv[1] if len(sys.argv) > 1 else "cbom_results.json"
    o = sys.argv[2] if len(sys.argv) > 2 else "TLS_PQC_Posture_Score.xlsx"
    generate_risk_report(j, o)
